import builtins
from typing import Counter
from faker import Faker
from array import *
import random
from openpyxl import Workbook, load_workbook

# For testing positive scenario of registration and loging in
wb = load_workbook(filename="RESOURCES/TestData.xlsx")
ws = wb.active
def not_registered_user():
    for i in range(2, 101):
        while(ws.cell(row=i, column=6).value ==None):
            row_id=i
            return row_id
            break

row_id=not_registered_user()
        
def new_user_data_username():           
    username = ws.cell(row=row_id, column=1).value
    return username
def new_user_data_password():
    user_password = ws.cell(row=row_id, column=2).value
    return user_password
def new_user_data_first_name():
    first_name = ws.cell(row=row_id, column=3).value
    return first_name
def new_user_data_last_name():
    last_name = ws.cell(row=row_id, column=4).value
    return last_name
def new_user_phone_number():
    phone_number = ws.cell(row=row_id, column=5).value
    return  phone_number
    
def add_registered():
    ws.cell(row=row_id, column=6).value = "Yes"
    wb.save("RESOURCES/TestData.xlsx")

# For testing negative scenario of registration and loging in

def registered_user():
    for i in range(2, 101):
        while(ws.cell(row=i, column=6).value =="Yes"):
            row_id_neg=i
            return row_id_neg
            break

row_id_neg=registered_user()


def registered_username():           
    username = ws.cell(row=row_id_neg, column=1).value
    return username

def user_password():
    user_password = ws.cell(row=row_id_neg, column=2).value
    return user_password
def registered_first_name():
    first_name = ws.cell(row=row_id_neg, column=3).value
    return first_name
def registered_last_name():
    last_name = ws.cell(row=row_id_neg, column=4).value
    return last_name
def registered_phone_number():
    phone_number = ws.cell(row=row_id_neg, column=5).value
    return  phone_number
    
def add_registered():
    ws.cell(row=row_id, column=6).value = "Yes"
    wb.save("RESOURCES/TestData.xlsx")

# randomly chosen registered user row from xlsx file

def registered_user_api():
    registered_user_array=array("I",[])
    for i in range(2, 101):
        if(ws.cell(row=i, column=6).value =="Yes"):
            registered_user_array.append(i)
    return random.choice(registered_user_array)

    
row_id_api=registered_user_api()

def registered_username_api():           
    username = ws.cell(row=row_id_api, column=1).value
    return username
    


def registered_user_password_api():
    user_password = ws.cell(row=row_id_api, column=2).value
    return user_password

def registered_user_fierst_name_api():
    first_name = ws.cell(row=row_id_api, column=3).value
    return first_name 

def registered_user_fierst_second_name_api():
    second_name = ws.cell(row=row_id_api, column=4).value
    return second_name

def registered_user_mobile_phone_api():
    mobile_phone = ws.cell(row=row_id_api, column=5).value
    return mobile_phone




#randomly chosen not registered user row from xlsx file
def not_registered_user_api():
    not_registered_user_array=array("I",[])
    for i in range(2, 101):
        if(ws.cell(row=i, column=6).value ==None):
            not_registered_user_array.append(i)
    # return registered_user_array
    return random.choice(not_registered_user_array)
    
row_id_api_neg=not_registered_user_api()

def not_registered_username_api():           
    username = ws.cell(row=row_id_api_neg, column=1).value
    return username


def not_user_password_api():
    user_password = ws.cell(row=row_id_api_neg, column=2).value
    return user_password


# collect all registered users
def collect_registered_users_xlsx():
    registered_user_array_xlsx=[]
    for i in range(2, 101):
        if(ws.cell(row=i, column=6).value =="Yes"):
            username=ws.cell(row=i, column=1).value
            registered_user_array_xlsx.append(username)
    registered_user_array_xlsx=str(registered_user_array_xlsx)
    return registered_user_array_xlsx


def invalid_token():
    fake_data= Faker()
    invalid_token=fake_data.bothify(text="??????##???????#")
    return invalid_token

