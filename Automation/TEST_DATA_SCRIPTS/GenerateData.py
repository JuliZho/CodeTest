from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
fake_data = Faker()
fake_phone_number = Faker("fi_FI")

#Generating test data
def generate_data():
    ws.cell(row=1, column=1).value = "Username"
    ws.cell(row=1, column=2).value = "Password"
    ws.cell(row=1, column=3).value = "First name"
    ws.cell(row=1, column=4).value = "Last name"
    ws.cell(row=1, column=5).value = "Phone number"
    ws.cell(row=1, column=6).value = "Registered"
    for i in range(2, 101):
        for j in range(1, 6):
            ws.cell(row=i, column=1).value = fake_data.bothify(text="??????###")
            ws.cell(row=i, column=2).value = fake_data.bothify(text="??#?#?#?#?#")
            ws.cell(row=i, column=3).value = fake_data.first_name()
            ws.cell(row=i, column=4).value = fake_data.last_name()
            ws.cell(row=i, column=5).value = fake_phone_number.phone_number()
    wb.save("RESOURCES/TestData.xlsx")
     
generate_data()